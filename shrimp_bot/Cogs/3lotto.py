from discord.ext import commands
from random import sample
from supporter import *
from data import prefix
import openpyxl


xl2 = openpyxl.load_workbook("user_money.xlsx")
sheet2 = xl2.active


class lotto(commands.Cog, name="로또 명령어"):
    def __init__(self, client):
        self.client = client

    @commands.command(name="로또구매", aliases=[ "ㄹㄸㄱㅁ", "로또생성", "로또", "ㄹㄸ"])
    async def make_lotto(self, ctx, lotto_n1:str=None, lotto_n2:str=None, lotto_n3:str=None, lotto_n4:str=None, lotto_n5:str=None, lotto_n6:str=None):
        row = await search_row("user_money.xlsx", ctx.author.id)
        if row == None:
            await ctx.reply(f"도박 관련 기능 명령어를 사용하려면\n`{prefix}사용자등록` 명령어로 등록을\n먼저 해주세요!")
            return
        if sheet2.cell(row=row, column=6).value != None:
            await ctx.reply("이미 로또를 구매했어요!")
            return
        lotto = []
        if ctx.message.content == f"{prefix}로또구매 자동" or ctx.message.content == f"{prefix}로또구매":
            lotto_nums = list(range(1, 46))
            lotto = sample(lotto_nums, 6)
            lotto.sort()
        else:
            try:
                n_list = [lotto_n1, lotto_n2, lotto_n3, lotto_n4, lotto_n5, lotto_n6]
                for n in n_list:
                    if n in lotto:
                        await ctx.reply("중복된 값은 들어올 수 없어요!")
                        return
                    lotto.append(int(n))
                lotto.sort()
            except ValueError:
                await ctx.reply("숫자만 입력해주세요!")
                return
            except:
                await ctx.reply("값을 정확히 입력해주세요!")
                return

            for n in lotto:
                if n < 1 or n > 45:
                    await ctx.reply("수는 1~45까지의 수만 들어올 수 있어요!")
                    return

        sheet2.cell(row=row, column=6).value = f"{lotto[0]} {lotto[1]} {lotto[2]} {lotto[3]} {lotto[4]} {lotto[5]}"
        xl2.save("user_money.xlsx")
        await ctx.reply("로또 구매 완료!")
        await ctx.reply(f"{ctx.author}님의 로또번호: {lotto[0]} {lotto[1]} {lotto[2]} {lotto[3]} {lotto[4]} {lotto[5]}")
        return


    @commands.command(name="보너스번호", aliases=[ "보너스", "ㅂㄴㅅㅂㅎ", "ㅂㄴㅅ" ])
    async def bonus_num(self, ctx, bonus_num):
        row = await search_row("user_money.xlsx", ctx.author.id)
        if row == None:
            await ctx.reply(f"도박 관련 기능 명령어를 사용하려면\n`{prefix}사용자등록` 명령어로 등록을\n먼저 해주세요!")
            return
        if sheet2.cell(row=row, column=6).value == None:
            await ctx.reply("이 명령어는 로또를 먼저 구매하고 나서 사용해주세요!")
            return
        try:
            if bonus_num < 1 or bonus_num > 45:
                await ctx.reply("수는 1~45까지의 수만 들어올 수 있어요!")
                return
        except:
            await ctx.reply(f"값을 정확히 입력해주세요!\n예시: {prefix}보너스번호 7")
            return
        sheet2.cell(row=row, column=7).value = f"{bonus_num}"
        xl2.save("user_money.xlsx")
        await ctx.reply(f"<@!{ctx.author.id}>님의 보너스 번호를 {bonus_num}으로 설정했어요!")
        return
    

    @commands.command(name="로또확인", alias="ㄹㄸㅎㅇ")
    async def check_lotto(self, ctx):
        row = await search_row("user_money.xlsx", ctx.author.id)
        if row == None:
            await ctx.reply(f"도박 관련 기능 명령어를 사용하려면\n`{prefix}사용자등록` 명령어로 등록을\n먼저 해주세요!")
            return
        count = 0
        prize = 0
        add = 0
        try:
            lotto = sheet2.cell(row=row, column=6).value.split(" ")
            if lotto == None:
                await ctx.reply("아직 로또를 구매하지 않았어요!")
                return
            bonus_num = int(sheet2.cell(row=row, column =7).value)
        except TypeError:
            bonus_num = "없음"
        except ValueError:
            bonus_num = "없음"
        except AttributeError:
            await ctx.reply("아직 로또를 구매하지 않았어요!")
            return

        win = sheet2.cell(row=2, column=8).value.split(" ")
        user_money = int(sheet2.cell(row=row, column=2).value)
        for n in lotto:
            if int(n) in win:
                count += 1
        if count == 6:
            prize = 1
            add = 7777777777777777
        elif count == 5:
            if bonus_num in win:
                prize = 2
                add = 777777777777
            else:
                prize = 3
                add = 77777777
        elif count == 4:
            prize = 4
            add = 7777777
        else:
            await ctx.reply(f"{ctx.author}님의 로또번호: {lotto[0]} {lotto[1]} {lotto[2]} {lotto[3]} {lotto[4]} {lotto[5]}\n보너스번호: {bonus_num}")
            await ctx.reply("꽝에 당첨되었어요........")
            sheet2.cell(row=row, column=6).value = None
            sheet2.cell(row=row, column=7).value = None
            xl2.save("user_money.xlsx")
            return
        await ctx.reply(f"{ctx.author}님의 로또번호: {lotto[0]} {lotto[1]} {lotto[2]} {lotto[3]} {lotto[4]} {lotto[5]}\n보너스번호: {bonus_num}")
        sheet2.cell(row=row, column=2).value = f"{user_money + add}"
        add = cal(add)
        await ctx.reply(f"{prize}등 복권에 당첨되었어요! 보상을 지급할게요!!\n`+ {add}`")
        if prize == 1 or prize == 2:
            lotto_nums = list(range(1, 46))
            win = sample(lotto_nums, 6)
            win.sort()
            sheet2.cell(row=2, column=7).value = f"{win[0]} {win[1]} {win[2]} {win[3]} {win[4]} {win[5]}"
            await ctx.reply(f"{prize}등에 당첨되어서 당첨번호가 재설정 되었어요!")
        sheet2.cell(row=row, column=6).value = None
        sheet2.cell(row=row, column =7).value = None
        xl2.save("user_money.xlsx")
        return




def setup(client):
    client.add_cog(lotto(client))
