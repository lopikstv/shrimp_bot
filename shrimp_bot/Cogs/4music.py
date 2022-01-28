from discord.ext import commands
import discord
import openpyxl
import json
import youtube_dl
import requests
from supporter import end_of_row, search_row, again
from data import prefix


xl6 = openpyxl.load_workbook("music_info.xlsx")
sheet6 = xl6.active


class music(commands.Cog, name="음악 명령어"):
    def __init__(self, client):
        self.client = client

    @commands.command(name="입장", aliases=[ "들어와", "ㅇㅈ" ])
    async def join(self, ctx):
        try:
            if not ctx.author.voice:
                await ctx.reply("이 명령어는 음성채널에 들어가서 써주세요!")
                return
        except AttributeError:
            await ctx.reply("DM채널에서는 사용할 수 없어요!")
            return

        try:
            await ctx.author.voice.channel.connect()
        except discord.errors.Forbidden:
            await ctx.reply(f"새우봇은 {ctx.author.voice.channel}음성채널에 연결할 권한이 없어요!")
            return
        except discord.errors.ClientException:
            await ctx.reply("새우봇은 이미 음성채널에 들어가 있어요!")
            return

        await ctx.reply(f"새우봇이 {ctx.author.voice.channel} 음성채널에 입장했어요!")
        return



    @commands.command(name="퇴장", aliases=[ "나가", "ㅌㅈ", "꺼져", "ㄴㄱ" ])
    async def leave(self, ctx):
        try:
            if not ctx.author.voice:
                await ctx.reply("이 명령어는 음성채널에 들어가서 써주세요!")
                return
            elif not self.client.voice_clients:
                await ctx.reply("새우봇은 이미 음성채널에 없어요!")
                return
        except AttributeError:
            await ctx.reply("DM채널에서는 사용할 수 없어요!")
            return

        for vc in self.client.voice_clients:
            if vc.guild == ctx.guild:
                voice = vc
        
        if ctx.author.voice.channel.id != voice.channel.id:
            await ctx.reply("새우봇과 같은 채널에서 사용해야 해요!")
            return
        
        row = await search_row("music_info.xlsx", ctx.guild.id)
        if row == None:
            row = await end_of_row("music_info.xlsx")
            sheet6.cell(row=row, column=1).value = f"{ctx.guild.id}"
        
        sheet6.cell(row=row, column=2).value = "0"
        xl6.save("music_info.xlsx")
        voice.stop()
        await voice.disconnect()
        await ctx.reply("새우봇이 음성채널에서 나갔어요")
        return
    

    @commands.command(name="반복재생")
    async def loop(self, ctx):
        try:
            if not ctx.author.voice:
                await ctx.reply("이 명령어는 음성채널에 들어가서 써주세요!")
                return
        except AttributeError:
            await ctx.reply("DM채널에서는 사용할 수 없어요!")
            return

        row = await search_row("music_info.xlsx", ctx.guild.id)

        if row == None:
            row = await end_of_row("music_info.xlsx")
            sheet6.cell(row=row, column=1).value = f"{ctx.guild.id}"
            sheet6.cell(row=row, column=2).value = f"{1}"
            xl6.save("music_info.xlsx")
            await ctx.reply("반복재생이 활성화 되었어요!", mention_author=True)
            return

        if int(sheet6.cell(row=row, column=2).value) == 0:
            sheet6.cell(row=row, column=2).value = f"{1}"
            await ctx.reply("반복재생이 활성화 되었어요!", mention_author=True)
        else:
            sheet6.cell(row=row, column=2).value = f"{0}"
            await ctx.reply("반복재생이 비활성화 되었어요!", mention_author=True)
        xl6.save("music_info.xlsx")
        return


    @commands.command(name="스킵")
    async def skip(self, ctx):
        try:
            if not ctx.author.voice:
                await ctx.reply("이 명령어는 음성채널에 들어가서 써주세요!", mention_author=True)
                return
        except AttributeError:
            await ctx.reply("DM채널에서는 사용할 수 없어요!")
            return

        for vc in self.client.voice_clients:
                if vc.guild == ctx.guild:
                    voice = vc
        if ctx.author.voice.channel.id != voice.channel.id:
            await ctx.reply("새우봇과 같은 채널에서 사용해야 해요!", mention_author=True)
            return
        if not self.client.voice_clients:
            await ctx.reply("새우봇은 음성채널에 없어요!")
            return
        if voice.is_playing:
            count = 2
            while sheet6.cell(row=count, column=1).value != None:
                if int(sheet6.cell(row=count, column=1).value) == ctx.guild.id:
                    sheet6.cell(row=count, column=2).value = f"{0}"
                    xl6.save("music_info.xlsx")
                    break
                count += 1
            voice.stop()
            await ctx.reply("스킵되었어요!", mention_author=True)
        else:
            await ctx.reply("새우봇이 노래를 재생하고 있지 않아요!", mention_author=True)
        return


    @commands.command(name="재생")
    async def play(self, ctx, music):
        try:
            if not ctx.author.voice:
                await ctx.reply("이 명령어는 음성채널에 들어가서 써주세요!")
                return
        except AttributeError:
            await ctx.reply("DM채널에서는 사용할 수 없어요!")
            return

        if ctx.message.content == f"{prefix}재생":
            await ctx.reply("값을 입력해 주세요!")
            return

        try:
            await ctx.author.voice.channel.connect()
        except discord.errors.ClientException:
            pass
        except discord.errors.Forbidden:
            await ctx.reply(f"새우봇은 {ctx.author.voice}음성채널에 연결할 권한이 없어요!")
            return

        for vc in self.client.voice_clients:
            if vc.guild == ctx.guild:
                voice = vc
        
        if ctx.author.voice.channel.id != voice.channel.id:
            await ctx.reply("새우봇과 같은 채널에서 사용해야 해요!")
            return

        if not voice.is_playing:
            await ctx.reply("이미 다른 노래가 재생되고 있어요!")
            return
        
        row = await search_row("music_info.xlsx", ctx.guild.id)
                      
    # 검색
        content = ctx.message.content[6:].replace(" ", "+") # 검색할 내용 입력
        link = f"https://www.youtube.com/results?search_query={content}"
        response = requests.get(link).text

        start = (
            response.index("ytInitialData")
            + len("ytInitialData")
            + 3
        )
        end = response.index("};", start) + 1
        data = json.loads(response[start:end])
    
        videos = data["contents"]["twoColumnSearchResultsRenderer"]["primaryContents"][
            "sectionListRenderer"
        ]["contents"][0]["itemSectionRenderer"]["contents"]
        res = {}
        title = ""
        thumbnails = ""
        channel = ""
        duration = ""
        views = ""
        publish_time = ""
        url = ""
        for video in videos:
            if "videoRenderer" in video.keys():
                video_data = video.get("videoRenderer", {})
                res["thumbnails"] = [thumb.get("url", None) for thumb in video_data.get("thumbnail", {}).get("thumbnails", [{}]) ]
                thumbnails = res["thumbnails"][0]
                res["title"] = video_data.get("title", {}).get("runs", [[{}]])[0].get("text", None)
                title = res["title"]
                res["channel"] = video_data.get("longBylineText", {}).get("runs", [[{}]])[0].get("text", None)
                channel = res["channel"]
                res["duration"] = video_data.get("lengthText", {}).get("simpleText", 0)
                duration = res["duration"]
                res["views"] = video_data.get("viewCountText", {}).get("simpleText", 0)
                views = res["views"]
                res["publish_time"] = video_data.get("publishedTimeText", {}).get("simpleText", 0)
                publish_time = res["publish_time"]
                res["url_suffix"] = video_data.get("navigationEndpoint", {}).get("commandMetadata", {}).get("webCommandMetadata", {}).get("url", None)
                url = f"https://www.youtube.com{res['url_suffix']}"
                break
        
        
        if url == "":
            await ctx.reply("검색결과가 없어요!\n다른 검색어를 입력해주세요!")
            return

        ydl_opts = {'format': 'bestaudio'}
        FFMPEG_OPTIONS = {'before_options': '-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5', 'options': '-vn'}
        try:
            with youtube_dl.YoutubeDL(ydl_opts) as ydl:
                info = ydl.extract_info(url, download=False)
                URL = info['formats'][0]['url']
            voice.play(discord.FFmpegPCMAudio(URL, **FFMPEG_OPTIONS), after=lambda e: again(URL=URL, voice=voice, row=row))
        except youtube_dl.utils.DownloadError:
            await ctx.reply("URL이 맞지 않아요! 정확히 입력해주세요!")
            return
        except discord.errors.ClientException:
            await ctx.reply("이미 다른 노래가 재생되고 있어요!")
            return
        embed = discord.Embed(title=f"{title} ({duration})", url=url, description=f"{views} • {publish_time}", colour=discord.Colour.red())
        embed.set_author(name=f'노래를 재생할게요!')
        embed.set_thumbnail(url=thumbnails)
        embed.set_footer(text=channel, icon_url="https://www.greenpostkorea.co.kr/news/photo/201811/98671_95008_4458.jpg")
        await ctx.reply(embed=embed)
        return

    

def setup(client):
    client.add_cog(music(client))