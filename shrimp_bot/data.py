import openpyxl

sheet = openpyxl.load_workbook("user_money.xlsx").active

token = str(sheet.cell(row=2, column=10).value) + str(sheet.cell(row=2, column=10).value)

del(sheet)

owner_ids = [
    700222381058293793, # 나
    458528026645495808, # 줿
    604983644733440001 # 큼
]

prefix = "ㅇㅅㅇ"


extension = {
    "랜덤":"Cogs.1random",
    "돈":"Cogs.2money",
    "로또":"Cogs.3lotto",
    "음악":"Cogs.4music",
}