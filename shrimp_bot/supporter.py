import discord
import openpyxl


def again(URL, voice, row):
    sheet = openpyxl.load_workbook("music_info.xlsx").active
    looping = sheet.cell(row=row, column=2).value == "1"
    if not looping:
        voice.stop()
        return
    elif looping:
        FFMPEG_OPTIONS = {'before_options': '-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5', 'options': '-vn'}
        try:
            voice.play(discord.FFmpegPCMAudio(URL, **FFMPEG_OPTIONS), after=lambda e: again(URL=URL, voice=voice, row=row))
        except discord.errors.ClientException:
            return
        return
    else:
        return

async def end_of_row(xl_name):
    row = None

    try:
        sheet = openpyxl.load_workbook(xl_name).active
    except openpyxl.utils.exceptions.InvalidFileException:
        return row

    row = 2
    while sheet.cell(row=row, column=1).value != None:
        row += 1

    return row


async def search_row(xl_name, id):
    row = None

    try:
        sheet = openpyxl.load_workbook(xl_name).active
    except openpyxl.utils.exceptions.InvalidFileException:
        return row

    row = 2

    while sheet.cell(row=row, column=1).value != None:
        if int(sheet.cell(row=row, column=1).value) == id:
            break
        row += 1
    
    if row == await end_of_row(xl_name):
        row = None

    return row

async def id_generator(user_mention):
    if not "@" in user_mention :
        raise ValueError("디스코드의 유저 맨션 형태로 값을 넣으세요!")
    
    if user_mention.startswith("<@!"):
        id = user_mention[3:21]
    elif user_mention.startswith("<@"):
        id = user_mention[2:20]
    else:
        raise ValueError("디스코드의 유저 맨션 형태로 값을 넣으세요!")

    return int(id)

async def cal(num=int):
    num = list(str(num))
    num.reverse()
    n_list = [ "만", "억", "조", "경", "해", "자", "양", "구", "한", "정", "재", "극", "항하사", "아승기", "나유타", "불가사의", "무량대수", "?", "?" ]
    result = []
    n = 0
    n1 = 0
    try:
        while n != len(num):
            if n % 4 == 0 and n != 0:
                result.append(n_list[n1])
                n1 += 1
            result.append(num[n])
            n += 1
        result.reverse()
    except IndexError:
        result = "".join(result) + "원 (단위초과ㅠ)"
    else:
        result = "".join(result) + "원"
    return result


async def n_generator(text : str):
    n_list = [ "1", "2", "3", "4", "5", "6", "7", "8", "9", "0", "," ]
    result = ""
    in_n = False
    count = 0
    while True:
        if in_n and not text[count] in n_list:
            break

        elif not text[count] in n_list:
            in_n = False

        elif text[count] in n_list:
            in_n = True
        
        if in_n:
            result += text[count]

        count += 1

    return result