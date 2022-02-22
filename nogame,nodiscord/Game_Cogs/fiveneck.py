from discord.ext import commands
import discord
from discord_components import ButtonStyle, Button, DiscordComponents
import openpyxl
import asyncio
from data import game_playing
from func import id_generator, search_row, now_time


xl = openpyxl.load_workbook("user_data.xlsx")
sheet = xl.active
game = "오목"

fivn_field = {}
fivn_player = []


class fiveneck(commands.Cog, name="오목"):
    def __init__(self, client):
        self.client = client
    

    @commands.command(name="오목", alias="ㅇㅁ")
    async def fiveneck(self, ctx, author : str):
        if ctx.author.id in game_playing:
            await ctx.reply("이미 게임을 플레이중이에요\n동시에 여러가지 게임은 할 수 없어요")
            return
        
        author_id = await id_generator(author)
        author_row = await search_row("user_data.xlsx", author_id)
        author_user = self.client.get_user(author_id)



        if author_row == None:
            await ctx.reply(f"{author_user}님은 No game, No discord의 플레이어가 아니에요!")
            return
        
        elif ctx.guild.get_member(author_id) == None: # author이 같은 길드에 없는 경우
            await ctx.reply(f"{author_user.name}님은 같은 서버에 속해있지 않아요")
            return
        
        elif sheet.cell(row=author_row, column=5).value == "0":
            await ctx.reply(f"{author_user.name}님은 지금 모든 게임신청을 받지 않고 있어요")
            return

        elif author_id in game_playing:
            await ctx.reply(f"{author_user.name}님은 이미 게임을 하고 있어요")
            return

        embed = discord.Embed(title="새로운 게임 신청!", description=f"보낸 사람: {ctx.author.name}\n길드: {ctx.guild.name}\n게임: {game}")
        embed.set_footer(text=await now_time)

        components = [
            Button(style=ButtonStyle.green, label="수락"),
            Button(style=ButtonStyle.red, label="거절")
        ]
        msg = await author_user.send(embed=embed, components=components)
        message = await ctx.send(f"{author_user.name}님에게 초대장을 보냈습니다\n15초 안에 수락을 받지 못하면 자동으로 거절됩니다")
        try:
            ddb = DiscordComponents(self.client)
            res = await ddb.wait_for("button_click", timeout=15, user=author_user, message=msg)
            if res.component.label == "거절":
                embed.title = "거절되었습니다"
                await msg.edit(embed=embed)
                await message.edit(content=f"<@!{ctx.author.id}>상대가 초대를 거절했습니다")
                return
        except asyncio.TimeoutError:
            embed.title = "시간 초과로 인해 거절되었습니다"
            await msg.edit(embed=embed)
            await message.edit(content=f"<@!{ctx.author.id}>시간 초과로 인해 거절되었습니다")
            return
        embed.title = "수락되었습니다"
        await msg.edit(embed=embed)
        await message.edit(content=f"<@!{ctx.author.id}>상대가 초대를 수락했어요!\n자, 게임을 시작하죠!")
        await asyncio.sleep(3)
        game_playing.append(ctx.author.id)
        game_playing.append(author_user.id)
        n1, n2 = int(ctx.author.id), int(author_user.id)
        if n1 > n2:
            n1, n2 = n2, n1
        fivn_player.append([n1, n2])
        fivn_field[[ n1, n2 ]] = []
        n = 1
        while n < 19 * 19:
            fivn_field[[ n1, n2 ]] = fivn_field[[ n1, n2 ]].append(0)
            n += 1


    @commands.command(name="오오목", alias="ㅇㅇㅁ")
    async def fivefiveneck(self, ctx, row : str, column : str):
        sheet.cell(row=await search_row("user_data.xlsx", ctx.author.id))


def setup(client):
    client.add_cog(fiveneck(client))