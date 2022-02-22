from discord.ext import commands
import discord
from discord_components import Button, ButtonStyle, DiscordComponents
import asyncio
import openpyxl
from func import end_of_row, search_row, now_time
from data import user_list


xl = openpyxl.load_workbook("user_data.xlsx")
sheet = xl.active


class base(commands.Cog, name="기본 명령어"):
    def __init__(self, client):
        self.client = client
    
    @commands.command(name="가입")
    async def sign_up(self, ctx):
        if ctx.author.id in user_list:
            await ctx.reply(f"{ctx.author}님은 이미 등록되어 있어요!")
            return
        try:
            embed = discord.Embed(title="No game, No discord\n이용 약관 동의", description='''동의 버튼을 누르면 앞으로 봇을 사용할 수 있습니다.
            그리고 봇을 사용할 때 사용자 식별을 위하여 
            사용자 고유 ID를 수집 및 저장하게 됩니다. 동의 하시겠습니까?
            (10초 안에 버튼을 눌러주세요)''', colour=discord.Colour.orange())
            msg = await ctx.send(embed=embed, 
            components=[
                Button(style=ButtonStyle.green, label="동의"),
                Button(style=ButtonStyle.red, label="취소")
            ])

            try:
                ddb = DiscordComponents(self.client)
                res = await ddb.wait_for("button_click", timeout=10, user=ctx.author)
                await msg.delete()
                if res.component.label == "취소":
                    msg = await ctx.send("취소되었습니다\n(이 메시지는 5초 후에 사라집니다)")
                    await asyncio.sleep(5)
                    await msg.delete()
                    return
            except asyncio.TimeoutError:
                await msg.delete()
                msg = await ctx.send("시간 초과로 인해 취소되었습니다\n(이 메시지는 5초 후에 사라집니다)")
                await asyncio.sleep(5)
                await msg.delete()
                return
            row = await end_of_row("user_money.xlsx")
            sheet.cell(row=row, column=1).value = f"{ctx.author.id}"
            sheet.cell(row=row, column=2).value = "100000"
            sheet.cell(row=row, column=3).value = "0"
            sheet.cell(row=row, column=4).value = "0"
            sheet.cell(row=row, column=5).value = "0"
            user_list.append(ctx.author.id)
            await ctx.send(f'사용자에 <@!{ctx.author.id}>님이 등록되었어요\n자, 게임을 시작하죠!')
            xl.save("user_data.xlsx")
        except PermissionError:
            await ctx.send("개발자가 사용자 정보 파일을 열고 있어요\n잠시 기다려주세요")
        return
    






def setup(client):
    client.add_cog(base(client))